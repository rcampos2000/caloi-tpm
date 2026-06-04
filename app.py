"""
Sistema de Controle de Manutenção - Caloi TPM v2.0
Desenvolvido para gestão de manutenção industrial baseada em TPM
"""

from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, session
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from functools import wraps
import os
import json
import base64
import hashlib
import time
from pathlib import Path
import shutil

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'caloi-tpm-manutencao-2024-secret')

# ============================================================
# CAMINHOS DO SISTEMA
# DATA_DIR: em producao na nuvem, aponta para o volume persistente
#           em desenvolvimento local, usa a pasta do app
# ============================================================
APP_DIR  = Path(__file__).parent
DATA_DIR = Path(os.environ.get('DATA_DIR', str(APP_DIR)))
DB_FILE         = DATA_DIR / "Manutencao_TPM.xlsx"
PLANO_ACAO_FILE = DATA_DIR / "Plano_de_Acao.xlsx"
COLAB_FILE      = DATA_DIR / "Colaboradores.xlsx"
ASSINATURAS_DIR = DATA_DIR / "assinaturas"
BACKUP_DIR      = DATA_DIR / "backups"
CONFIG_DIR      = DATA_DIR / "config"
FOTOS_DIR       = DATA_DIR / "fotos"
USUARIOS_FILE   = CONFIG_DIR / "usuarios.json"
LISTAS_FILE     = CONFIG_DIR / "listas.json"
EQUIPAMENTOS_FILE = CONFIG_DIR / "equipamentos.json"
PRODUCAO_FILE     = CONFIG_DIR / "producao.json"
ESTOQUE_FILE      = CONFIG_DIR / "estoque.json"

# ============================================================
# DADOS PADRÃO (usados na primeira inicialização)
# ============================================================
TECNICOS_PADRAO = [
    "Anderson Silva", "Carlos Eduardo", "Diego Ferreira",
    "Eduardo Santos", "Felipe Oliveira", "Gabriel Rodrigues",
    "Henrique Costa", "Ivan Martins", "João Paulo",
    "Lucas Pereira", "Marcos Alves", "Nelson Souza",
    "Pedro Henrique", "Rafael Lima", "Rodrigo Mendes",
    "Thiago Barbosa", "Wagner Nunes",
    "Outro (informar no campo observações)"
]

SETORES_PADRAO = [
    "Preparação Aço", "Preparação Alumínio", "Solda Aço",
    "Solda Alumínio", "Tratamento Térmico", "Pintura",
    "Linha de Montagem", "Aro", "Rodas", "Outros"
]

MOTIVOS_PARADA = [
    "Selecione o motivo...",
    "Falha Mecânica - Quebra de componente",
    "Falha Mecânica - Desgaste de peça",
    "Falha Mecânica - Desalinhamento",
    "Falha Elétrica - Queima de motor",
    "Falha Elétrica - Problema no painel",
    "Falha Elétrica - Sensor defeituoso",
    "Falha Eletrônica - CLP/Inversor",
    "Falha Eletrônica - IHM/Display",
    "Falha Hidráulica - Vazamento",
    "Falha Hidráulica - Bomba",
    "Falha Pneumática - Vazamento de ar",
    "Falha Pneumática - Cilindro",
    "Falta de Material / Insumo",
    "Setup / Troca de Ferramenta",
    "Troca de Molde / Gabarito",
    "Manutenção Preventiva Programada",
    "Lubrificação / Engrase",
    "Limpeza Geral / 5S",
    "Ajuste de Parâmetros / Regulagem",
    "Colisão / Acidente de Máquina",
    "Problema de Qualidade - Ajuste",
    "Falta de Energia Elétrica",
    "Problema de Rede / Sistema",
    "Outros (descrever em observações)"
]

SOLUCOES = [
    "Selecione a solução...",
    "Substituição de Peça / Componente",
    "Ajuste e Regulagem",
    "Lubrificação",
    "Limpeza e Higienização",
    "Reparo Elétrico - Troca de fusível",
    "Reparo Elétrico - Troca de motor",
    "Reparo Elétrico - Reaperto de conexões",
    "Reparo Eletrônico - Parametrização",
    "Reparo Eletrônico - Substituição de placa",
    "Reparo Mecânico - Solda",
    "Reparo Mecânico - Usinagem",
    "Reparo Hidráulico - Troca de vedação",
    "Reparo Pneumático - Troca de mangueira",
    "Troca de Correia / Corrente",
    "Troca de Rolamento",
    "Troca de Filtro",
    "Calibração / Zeragem",
    "Reset / Reinicialização do Sistema",
    "Aguardou normalização (energia/rede)",
    "Manutenção preventiva executada",
    "Intervenção do fabricante / terceiros",
    "Outros (descrever em observações)"
]

EQUIPAMENTOS_CATALOGO = {
    "EQ001": "Prensa Hidráulica #1 - Preparação Aço",
    "EQ002": "Prensa Hidráulica #2 - Preparação Aço",
    "EQ003": "Dobradeira CNC - Preparação Aço",
    "EQ004": "Guilhotina - Preparação Aço",
    "EQ005": "Prensa Alumínio #1 - Preparação Alumínio",
    "EQ006": "Extrusora - Preparação Alumínio",
    "EQ007": "Robô de Solda #1 - Solda Aço",
    "EQ008": "Robô de Solda #2 - Solda Aço",
    "EQ009": "Solda MIG Manual - Solda Aço",
    "EQ010": "Robô de Solda Alumínio #1 - Solda Alumínio",
    "EQ011": "Forno de Têmpera - Tratamento Térmico",
    "EQ012": "Forno de Revenimento - Tratamento Térmico",
    "EQ013": "Cabine de Pintura #1 - Pintura",
    "EQ014": "Cabine de Pintura #2 - Pintura",
    "EQ015": "Forno de Cura - Pintura",
    "EQ016": "Esteira Linha A - Linha de Montagem",
    "EQ017": "Esteira Linha B - Linha de Montagem",
    "EQ018": "Laminadora de Aro #1 - Aro",
    "EQ019": "Laminadora de Aro #2 - Aro",
    "EQ020": "Montadora de Rodas - Rodas",
}


# ============================================================
# FUNÇÕES DE USUÁRIOS E CONFIGURAÇÃO
# ============================================================
def hash_senha(senha):
    return hashlib.sha256(senha.encode('utf-8')).hexdigest()


def carregar_usuarios():
    if USUARIOS_FILE.exists():
        with open(USUARIOS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []


def salvar_usuarios(usuarios):
    CONFIG_DIR.mkdir(exist_ok=True)
    with open(USUARIOS_FILE, 'w', encoding='utf-8') as f:
        json.dump(usuarios, f, ensure_ascii=False, indent=2)


def carregar_listas():
    if LISTAS_FILE.exists():
        with open(LISTAS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {'tecnicos': TECNICOS_PADRAO, 'setores': SETORES_PADRAO}


def salvar_listas(listas):
    CONFIG_DIR.mkdir(exist_ok=True)
    with open(LISTAS_FILE, 'w', encoding='utf-8') as f:
        json.dump(listas, f, ensure_ascii=False, indent=2)


COLAB_BASE_HEADERS = ["Nome", "Função", "Setor", "Admissão",
                      "Férias Início 1", "Férias Fim 1", "Férias Início 2", "Férias Fim 2"]
COLAB_CAP_MAX = 10
COLAB_CAP_HEADERS = [h for i in range(1, COLAB_CAP_MAX + 1)
                     for h in (f"Capacitação {i}", f"Validade {i}")]
COLAB_HEADERS = COLAB_BASE_HEADERS + COLAB_CAP_HEADERS
FUNCOES_PADRAO = ["Mecânico", "Eletricista", "Ferramenteiro", "Serralheiro", "Auxiliar de Manutenção"]

def carregar_colaboradores():
    """Lê Colaboradores.xlsx -> lista de dicts (chaves = COLAB_HEADERS)."""
    if not COLAB_FILE.exists():
        return []
    try:
        wb = openpyxl.load_workbook(COLAB_FILE, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return []
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        out = []
        for r in rows[1:]:
            d = {headers[i]: ("" if i >= len(r) or r[i] is None else str(r[i]).strip())
                 for i in range(len(headers))}
            if d.get("Nome"):
                out.append(d)
        return out
    except Exception as e:
        print(f"[WARN] carregar_colaboradores: {e}")
        return []

def salvar_colaboradores(lista):
    """Grava a lista de colaboradores em Colaboradores.xlsx."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Colaboradores"
    ws.append(COLAB_HEADERS)
    for c in lista:
        ws.append([c.get(h, "") for h in COLAB_HEADERS])
    hf = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="1F3A8F")
    for cell in ws[1]:
        cell.font = hf
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")
    widths = [20, 22, 18, 12, 14, 14, 14, 14] + [18, 12] * COLAB_CAP_MAX
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    wb.save(COLAB_FILE)

def get_tecnicos():
    listas = carregar_listas()
    return ['Selecione o técnico...'] + listas.get('tecnicos', TECNICOS_PADRAO)


def get_setores():
    listas = carregar_listas()
    return listas.get('setores', SETORES_PADRAO)


def get_motivos():
    listas = carregar_listas()
    return listas.get('motivos', MOTIVOS_PARADA)


def get_solucoes():
    listas = carregar_listas()
    return listas.get('solucoes', SOLUCOES)


def carregar_equipamentos():
    if EQUIPAMENTOS_FILE.exists():
        with open(EQUIPAMENTOS_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    # Fallback: converter catálogo legado para o novo formato
    return [
        {
            'descricao': v.split(' - ')[0] if ' - ' in v else v,
            'tag': k,
            'area': v.split(' - ')[-1] if ' - ' in v else '',
            'fornecedor': ''
        }
        for k, v in EQUIPAMENTOS_CATALOGO.items()
    ]


def salvar_equipamentos(equipamentos):
    CONFIG_DIR.mkdir(exist_ok=True)
    with open(EQUIPAMENTOS_FILE, 'w', encoding='utf-8') as f:
        json.dump(equipamentos, f, ensure_ascii=False, indent=2)


def equipamentos_como_catalogo():
    """Retorna equipamentos no formato {TAG: 'Descrição - Área'} para compatibilidade."""
    eqs = carregar_equipamentos()
    return {
        eq['tag']: f"{eq['descricao']}{' - ' + eq['area'] if eq.get('area') else ''}"
        for eq in eqs if eq.get('tag')
    }


def equipamentos_meta():
    """Retorna {TAG: {descricao, area}} para preenchimento automático de setor no formulário."""
    eqs = carregar_equipamentos()
    return {
        eq['tag']: {'descricao': eq.get('descricao', ''), 'area': eq.get('area', eq.get('setor', ''))}
        for eq in eqs if eq.get('tag')
    }


def carregar_producao():
    """Retorna lista de config OEE por equipamento."""
    if PRODUCAO_FILE.exists():
        with open(PRODUCAO_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    # Inicializa vazia — o admin preenche depois
    return []


def salvar_producao(dados):
    CONFIG_DIR.mkdir(exist_ok=True)
    with open(PRODUCAO_FILE, 'w', encoding='utf-8') as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)


def get_status_por_equipamento():
    """
    Lê Manutencao_TPM.xlsx e retorna dict {TAG: último_status}.
    Cruza com equipamentos cadastrados para identificar pendentes.
    Retorna:
      { tag: 'Pendente' | 'Concluído' | 'Sem registros' }
    """
    equipamentos = carregar_equipamentos()
    resultado = {eq['tag']: 'Sem registros' for eq in equipamentos if eq.get('tag')}

    if not DB_FILE.exists():
        return resultado

    try:
        wb = openpyxl.load_workbook(DB_FILE, read_only=True)
        ws = wb["Registros"]
        headers = [cell.value for cell in ws[1]]
        # Índices das colunas que precisamos
        try:
            # Suporta ambos os nomes de coluna (compatibilidade)
            idx_cod   = next((headers.index(n) for n in ('Código do Equipamento', 'Código Equipamento') if n in headers), None)
            if idx_cod is None:
                wb.close()
                return resultado
            idx_stat  = headers.index('Status')
            idx_id    = headers.index('ID')
        except ValueError:
            wb.close()
            return resultado

        ultimo_por_tag = {}  # {tag: (id, status)}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(v is not None for v in row):
                continue
            tag    = str(row[idx_cod]).strip().upper() if row[idx_cod] else ''
            status = str(row[idx_stat]).strip() if row[idx_stat] else 'Concluído'
            rid    = row[idx_id] or 0
            if tag and (tag not in ultimo_por_tag or rid > ultimo_por_tag[tag][0]):
                ultimo_por_tag[tag] = (rid, status)

        wb.close()

        for tag, (_, status) in ultimo_por_tag.items():
            if tag in resultado:
                resultado[tag] = status
    except Exception:
        pass

    return resultado


def usuario_logado():
    uid = session.get('user_id')
    if not uid:
        return None
    for u in carregar_usuarios():
        if u['id'] == uid and u.get('ativo', True):
            return u
    return None


# ============================================================
# DECORADORES DE ACESSO
# ============================================================
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not usuario_logado():
            return redirect(url_for('login', next=request.path))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        u = usuario_logado()
        if not u:
            return redirect(url_for('login'))
        if u.get('perfil') != 'admin':
            return redirect(url_for('formulario'))
        return f(*args, **kwargs)
    return decorated


# ============================================================
# INICIALIZAÇÃO
# ============================================================
def criar_excel_plano_acao():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plano de Ação"

    headers = [
        "ID Registro", "Data/Hora", "Equipamento", "Código",
        "Setor", "Técnico", "Outros Problemas Encontrados",
        "Peças Necessárias", "Status"
    ]

    font_header = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    fill_header = PatternFill(start_color="7D4E00", end_color="B8860B", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin

    larguras = [12, 20, 28, 14, 20, 22, 50, 50, 14]
    for col, larg in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(col)].width = larg

    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(PLANO_ACAO_FILE)
    wb.close()
    print(f"  Plano de Ação criado: {PLANO_ACAO_FILE}")


def carregar_estoque():
    """Retorna dict {TAG: [lista de peças]} do estoque.json."""
    if ESTOQUE_FILE.exists():
        with open(ESTOQUE_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {}


def salvar_estoque(dados):
    CONFIG_DIR.mkdir(exist_ok=True)
    with open(ESTOQUE_FILE, 'w', encoding='utf-8') as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)


def salvar_fotos(fotos_b64, record_id):
    """Salva lista de imagens base64 em FOTOS_DIR. Retorna lista de nomes de arquivo."""
    FOTOS_DIR.mkdir(exist_ok=True)
    nomes = []
    for i, b64 in enumerate(fotos_b64[:3], 1):
        if not b64 or b64 == 'data:,':
            continue
        try:
            if ',' in b64:
                b64 = b64.split(',', 1)[1]
            dados = base64.b64decode(b64)
            nome = f"foto_{record_id:04d}_{i}_{datetime.now().strftime('%Y%m%d%H%M%S')}.jpg"
            caminho = FOTOS_DIR / nome
            with open(caminho, 'wb') as f:
                f.write(dados)
            nomes.append(nome)
        except Exception as e:
            print(f"  Aviso: erro ao salvar foto {i}: {e}")
    return nomes


def salvar_fotos_excel(record_id, fotos_b64):
    """Salva fotos como base64 na aba 'Fotos' do Excel principal (mesmo arquivo de OS)."""
    if not fotos_b64:
        return False
    max_tentativas = 3
    for tentativa in range(1, max_tentativas + 1):
        temp_path = DB_FILE.parent / f"_temp_fotos_{record_id}.xlsx"
        try:
            wb = openpyxl.load_workbook(DB_FILE)
            if "Fotos" not in wb.sheetnames:
                ws_fotos = wb.create_sheet("Fotos")
                ws_fotos.append(["ID", "Foto 1 (base64)", "Foto 2 (base64)", "Foto 3 (base64)"])
                ws_fotos.column_dimensions['A'].width = 8
                ws_fotos.column_dimensions['B'].width = 30
                ws_fotos.column_dimensions['C'].width = 30
                ws_fotos.column_dimensions['D'].width = 30
            else:
                ws_fotos = wb["Fotos"]
            row = [record_id]
            for b64 in fotos_b64[:3]:
                if b64 and ',' in b64:
                    b64 = b64.split(',', 1)[1]
                row.append(b64 or '')
            while len(row) < 4:
                row.append('')
            ws_fotos.append(row)
            wb.save(temp_path)
            wb.close()
            os.replace(temp_path, DB_FILE)
            return True
        except Exception as e:
            if temp_path.exists():
                try:
                    temp_path.unlink()
                except Exception:
                    pass
            if tentativa == max_tentativas:
                print(f"[ERRO] salvar_fotos_excel: {e}")
    return False


def init_database():
    APP_DIR.mkdir(parents=True, exist_ok=True)
    ASSINATURAS_DIR.mkdir(exist_ok=True)
    BACKUP_DIR.mkdir(exist_ok=True)
    CONFIG_DIR.mkdir(exist_ok=True)
    FOTOS_DIR.mkdir(exist_ok=True)

    if not DB_FILE.exists():
        criar_excel_banco_dados()
        print(f"  Banco de dados criado: {DB_FILE}")
    else:
        print(f"  Banco de dados existente: {DB_FILE}")

    if not PLANO_ACAO_FILE.exists():
        criar_excel_plano_acao()
    else:
        print(f"  Plano de Ação existente: {PLANO_ACAO_FILE}")


def init_config():
    CONFIG_DIR.mkdir(exist_ok=True)

    # Criar admin padrão se não houver usuários
    if not USUARIOS_FILE.exists() or not carregar_usuarios():
        usuarios = [{
            'id': 1,
            'username': 'admin',
            'nome': 'Administrador',
            'password_hash': hash_senha('admin123'),
            'perfil': 'admin',
            'ativo': True,
            'criado_em': datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        }]
        salvar_usuarios(usuarios)
        print("  ✓ Usuário padrão criado: admin / admin123")

    # Criar listas padrão se não existirem
    if not LISTAS_FILE.exists():
        salvar_listas({
            'tecnicos': TECNICOS_PADRAO,
            'setores': SETORES_PADRAO,
            'motivos': MOTIVOS_PARADA,
            'solucoes': SOLUCOES
        })
        print("  ✓ Listas de configuração criadas")
    else:
        # Adicionar chaves novas (motivos/solucoes) caso o arquivo já exista sem elas
        listas_existentes = carregar_listas()
        atualizado = False
        if 'motivos' not in listas_existentes:
            listas_existentes['motivos'] = MOTIVOS_PARADA
            atualizado = True
        if 'solucoes' not in listas_existentes:
            listas_existentes['solucoes'] = SOLUCOES
            atualizado = True
        if atualizado:
            salvar_listas(listas_existentes)
            print("  ✓ Listas de configuração atualizadas (motivos/solucoes adicionados)")

    # Criar producao.json vazio se não existir
    if not PRODUCAO_FILE.exists():
        salvar_producao([])
        print("  ✓ producao.json criado (vazio — configure em Configurações > Produção)")

    # Criar estoque.json vazio se não existir
    if not ESTOQUE_FILE.exists():
        salvar_estoque({})
        print("  ✓ estoque.json criado (vazio — configure em Configurações > Estoque)")

    # Sincronizar equipamentos.json: usa o arquivo bundled do repositório
    # quando o volume não tem o arquivo OU quando o bundled tem mais equipamentos
    # (permite atualizar a lista via deploy sem perder edições manuais feitas na web)
    BUNDLED_EQUIPAMENTOS = APP_DIR / "config" / "equipamentos.json"
    if BUNDLED_EQUIPAMENTOS.exists():
        try:
            with open(BUNDLED_EQUIPAMENTOS, 'r', encoding='utf-8') as _f:
                bundled_eqps = json.load(_f)
            if not EQUIPAMENTOS_FILE.exists():
                salvar_equipamentos(bundled_eqps)
                print(f"  ✓ Equipamentos carregados do bundled: {len(bundled_eqps)} equipamentos")
            else:
                volume_eqps = carregar_equipamentos()
                if len(bundled_eqps) > len(volume_eqps):
                    salvar_equipamentos(bundled_eqps)
                    print(f"  ✓ Equipamentos atualizados do bundled: {len(bundled_eqps)} (antes: {len(volume_eqps)})")
        except Exception as _e:
            print(f"  ⚠ Erro ao sincronizar equipamentos bundled: {_e}")
    elif not EQUIPAMENTOS_FILE.exists():
        # Fallback legado caso o bundled não exista
        equipamentos_padrao = [
            {
                'descricao': v.split(' - ')[0] if ' - ' in v else v,
                'tag': k,
                'area': v.split(' - ')[-1] if ' - ' in v else '',
                'fornecedor': ''
            }
            for k, v in EQUIPAMENTOS_CATALOGO.items()
        ]
        salvar_equipamentos(equipamentos_padrao)
        print("  ✓ Equipamentos migrados do catálogo legado")


def criar_excel_banco_dados():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registros"

    headers = [
        "ID", "Data/Hora Registro", "Data Ocorrência", "Equipamento",
        "Código do Equipamento", "Técnico Responsável", "Setor Produtivo",
        "Horário de Parada", "Horário de Liberação", "Total Horas Paradas",
        "Motivo da Parada", "Solução do Problema", "Observações",
        "Assinatura Técnico", "Nome Solicitante", "Liberação Solicitante", "Status",
        "Materiais Utilizados", "Fotos Registradas"
    ]

    font_header = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    fill_header = PatternFill(start_color="1B3A6B", end_color="1B3A6B", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin

    larguras = [6, 20, 15, 28, 16, 22, 20, 15, 15, 15, 30, 30, 25, 16, 20, 16, 12, 40, 14]
    for col, larg in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(col)].width = larg

    ws.row_dimensions[1].height = 35
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Aba Resumo
    ws2 = wb.create_sheet("Resumo por Setor")
    ws2['A1'] = "RESUMO DE MANUTENÇÃO POR SETOR"
    ws2['A1'].font = Font(bold=True, size=14, color="1B3A6B", name="Calibri")
    ws2['A1'].alignment = align_center
    ws2.merge_cells('A1:D1')

    for col, h in enumerate(["Setor", "Qtd. Paradas", "Total Horas", "Média Horas/Parada"], 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    for row, setor in enumerate(get_setores(), 3):
        ws2.cell(row=row, column=1, value=setor).alignment = align_center
        ws2.cell(row=row, column=2, value=f"=COUNTIF(Registros!G:G,A{row})").alignment = align_center
        ws2.cell(row=row, column=3, value=f"=SUMIF(Registros!G:G,A{row},Registros!J:J)").alignment = align_center
        ws2.cell(row=row, column=4, value=f"=IFERROR(C{row}/B{row},0)").alignment = align_center

    for col in [1, 2, 3, 4]:
        ws2.column_dimensions[get_column_letter(col)].width = 25

    # Aba Catálogo
    ws3 = wb.create_sheet("Catálogo Equipamentos")
    ws3['A1'] = "CATÁLOGO DE EQUIPAMENTOS"
    ws3['A1'].font = Font(bold=True, size=14, color="1B3A6B", name="Calibri")
    ws3['A1'].alignment = align_center
    ws3.merge_cells('A1:C1')

    for col, h in enumerate(["Código", "Descrição do Equipamento", "Setor"], 1):
        cell = ws3.cell(row=2, column=col, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center

    for row, (cod, desc) in enumerate(EQUIPAMENTOS_CATALOGO.items(), 3):
        ws3.cell(row=row, column=1, value=cod)
        ws3.cell(row=row, column=2, value=desc)
        setor = desc.split(" - ")[-1] if " - " in desc else ""
        ws3.cell(row=row, column=3, value=setor)

    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 40
    ws3.column_dimensions['C'].width = 25

    wb.save(DB_FILE)


def fazer_backup():
    if DB_FILE.exists():
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_file = BACKUP_DIR / f"Backup_{timestamp}_Manutencao_TPM.xlsx"
        shutil.copy2(DB_FILE, backup_file)
        backups = sorted(BACKUP_DIR.glob("Backup_*.xlsx"))
        if len(backups) > 10:
            for old in backups[:-10]:
                old.unlink()


# ============================================================
# ROTAS: LOGIN / LOGOUT
# ============================================================
@app.route('/login', methods=['GET', 'POST'])
def login():
    erro = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        senha = request.form.get('senha', '')
        senha_hash = hash_senha(senha)

        for u in carregar_usuarios():
            if u['username'] == username and u['password_hash'] == senha_hash and u.get('ativo', True):
                session['user_id'] = u['id']
                session['username'] = u['username']
                session['nome'] = u['nome']
                session['perfil'] = u['perfil']
                # Redireciona para a página que o usuário tentou acessar
                next_page = request.form.get('next') or request.args.get('next', '')
                # Segurança: só aceitar caminhos internos
                if next_page and next_page.startswith('/') and not next_page.startswith('//'):
                    return redirect(next_page)
                # Detectar mobile pelo User-Agent para redirecionar automaticamente
                ua = request.headers.get('User-Agent', '').lower()
                is_mobile = any(x in ua for x in ['android', 'iphone', 'ipad', 'mobile', 'phone'])
                if is_mobile:
                    return redirect(url_for('mobile_form'))
                return redirect(url_for('formulario'))

        erro = 'Usuário ou senha incorretos.'

    return render_template('login.html', erro=erro)


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ============================================================
# CONTEXT PROCESSOR — injeta 'usuario' em todos os templates
# ============================================================
@app.context_processor
def inject_usuario():
    return {'usuario': usuario_logado()}


# ============================================================
# ROTAS: INÍCIO, AJUDA E NAVEGAÇÃO GERAL
# ============================================================
@app.route('/')
def index():
    return redirect(url_for('inicio'))


@app.route('/inicio')
@login_required
def inicio():
    try:
        # Estatísticas rápidas para a tela inicial
        total_registros = 0
        total_horas = 0.0
        if DB_FILE.exists():
            wb = openpyxl.load_workbook(DB_FILE)
            ws = wb["Registros"]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(v is not None for v in row):
                    total_registros += 1
                    h = row[9]  # coluna "Total Horas Paradas"
                    if h and isinstance(h, (int, float)):
                        total_horas += h
            wb.close()
        listas = carregar_listas()
        return render_template(
            'inicio.html',
            hoje=datetime.now().strftime('%d/%m/%Y — %H:%M'),
            total_registros=total_registros,
            total_horas=round(total_horas, 1),
            total_tecnicos=len(listas.get('tecnicos', [])),
            total_setores=len(listas.get('setores', []))
        )
    except Exception as e:
        return render_template(
            'inicio.html',
            hoje=datetime.now().strftime('%d/%m/%Y — %H:%M'),
            total_registros=0, total_horas=0.0,
            total_tecnicos=0, total_setores=0
        )


@app.route('/ajuda')
@login_required
def ajuda():
    return render_template('ajuda.html')


@app.route('/formulario')
@login_required
def formulario():
    return render_template(
        'form.html',
        tecnicos=get_tecnicos(),
        setores=get_setores(),
        motivos=get_motivos(),
        solucoes=get_solucoes(),
        equipamentos=equipamentos_como_catalogo(),
        equipamentos_meta=equipamentos_meta(),
        tecnico_logado=session.get('nome', ''),
        data_hoje=datetime.now().strftime('%d/%m/%Y'),
        hora_atual=datetime.now().strftime('%H:%M')
    )


@app.route('/manifest.json')
def manifest():
    """Manifest PWA para instalação como app no celular."""
    from flask import send_from_directory
    return send_from_directory(APP_DIR / 'static', 'manifest.json',
                               mimetype='application/manifest+json')

@app.route('/mobile')
@login_required
def mobile_form():
    return render_template(
        'mobile_form.html',
        tecnicos=get_tecnicos(),
        setores=get_setores(),
        motivos=get_motivos(),
        solucoes=get_solucoes(),
        equipamentos=equipamentos_como_catalogo(),
        equipamentos_meta=equipamentos_meta(),
        tecnico_logado=session.get('nome', ''),
        data_hoje=datetime.now().strftime('%d/%m/%Y'),
        hora_atual=datetime.now().strftime('%H:%M'),
        session_user=session.get('usuario', '')
    )


@app.route('/api/estoque/<tag>')
@login_required
def api_estoque(tag):
    estoque = carregar_estoque()
    pecas = estoque.get(tag.upper().strip(), [])
    return jsonify({'tag': tag.upper(), 'pecas': pecas})


@app.route('/admin/estoque/salvar', methods=['POST'])
@admin_required
def admin_salvar_estoque():
    try:
        dados = request.get_json()
        estoque = dados.get('estoque', {})
        # Normalizar TAGs para maiúsculas
        estoque_norm = {k.upper().strip(): v for k, v in estoque.items() if k.strip()}
        salvar_estoque(estoque_norm)
        return jsonify({'success': True, 'message': 'Estoque salvo com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/api/equipamento/<codigo>')
@login_required
def buscar_equipamento(codigo):
    cod = codigo.upper().strip()
    # Busca primeiro no JSON configurado
    for eq in carregar_equipamentos():
        if eq.get('tag', '').upper() == cod:
            nome = eq['descricao']
            if eq.get('area'):
                nome += f" - {eq['area']}"
            return jsonify({'found': True, 'nome': nome, 'codigo': cod,
                            'descricao': eq.get('descricao', nome),
                            'area': eq.get('area', ''), 'setor': eq.get('setor', eq.get('area', '')),
                            'fornecedor': eq.get('fornecedor', '')})
    # Fallback no catálogo legado (compatibilidade)
    nome_legado = EQUIPAMENTOS_CATALOGO.get(cod)
    if nome_legado:
        return jsonify({'found': True, 'nome': nome_legado, 'codigo': cod})
    return jsonify({'found': False, 'codigo': cod})


def salvar_plano_acao(record_id, data, outros_problemas, pecas_necessarias):
    """Salva um item no Plano_de_Acao.xlsx usando gravação atômica."""
    max_tentativas = 5
    for tentativa in range(1, max_tentativas + 1):
        temp_path = PLANO_ACAO_FILE.parent / f"_temp_plano_{record_id}.xlsx"
        try:
            wb = openpyxl.load_workbook(PLANO_ACAO_FILE)
            ws = wb["Plano de Ação"]
            linha = ws.max_row + 1

            fill_color = "FFF8E1" if linha % 2 == 0 else "FFFFFF"
            row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            align_wrap = Alignment(vertical="top", wrap_text=True)
            align_center = Alignment(horizontal="center", vertical="center")
            border_thin = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )

            row_data = [
                record_id,
                datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
                data.get('equipamento_nome', ''),
                data.get('equipamento_codigo', ''),
                data.get('setor', ''),
                data.get('tecnico', ''),
                outros_problemas,
                pecas_necessarias,
                'Pendente'
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=linha, column=col, value=value)
                cell.fill = row_fill
                cell.border = border_thin
                cell.alignment = align_wrap if col in (7, 8) else align_center

            # Destaque visual na coluna Status (amarelo)
            status_cell = ws.cell(row=linha, column=9)
            status_cell.font = Font(bold=True, color="7D4E00")
            status_cell.fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

            ws.row_dimensions[linha].height = 40

            wb.save(temp_path)
            wb.close()
            os.replace(str(temp_path), str(PLANO_ACAO_FILE))
            print(f"  ✓ Plano de Ação #{record_id} salvo.")
            return

        except PermissionError:
            if temp_path.exists():
                try: temp_path.unlink()
                except: pass
            if tentativa < max_tentativas:
                time.sleep(1.5)
        except Exception as e:
            if temp_path.exists():
                try: temp_path.unlink()
                except: pass
            raise RuntimeError(f"Erro ao salvar Plano de Ação: {e}") from e


@app.route('/submit', methods=['POST'])
@login_required
def submit():
    try:
        data = request.get_json()
        total_horas = calcular_horas(data.get('horario_parada', ''), data.get('horario_liberacao', ''))
        record_id, next_row = obter_proximo_id()

        # Campos do Plano de Ação (opcionais)
        outros_problemas = data.get('outros_problemas', '').strip()
        pecas_necessarias = data.get('pecas_necessarias', '').strip()
        tem_plano_acao = bool(outros_problemas or pecas_necessarias)

        # Status: Pendente se há plano de ação, Concluído caso contrário
        status = 'Pendente' if tem_plano_acao else 'Concluído'

        sig_tecnico_path = salvar_assinatura(
            data.get('assinatura_tecnico', ''),
            f"tecnico_{record_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        )
        sig_solicitante_path = salvar_assinatura(
            data.get('assinatura_solicitante', ''),
            f"solicitante_{record_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        )

        # Fotos
        fotos_b64 = data.get('fotos', [])
        nomes_fotos = salvar_fotos(fotos_b64, record_id) if fotos_b64 else []

        # Materiais utilizados
        materiais = data.get('materiais_utilizados', [])
        materiais_str = ', '.join(materiais) if materiais else ''

        row_data = [
            record_id,
            datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
            data.get('data_ocorrencia', datetime.now().strftime('%d/%m/%Y')),
            data.get('equipamento_nome', ''),
            data.get('equipamento_codigo', ''),
            data.get('tecnico', ''),
            data.get('setor', ''),
            data.get('horario_parada', ''),
            data.get('horario_liberacao', ''),
            round(total_horas, 2) if total_horas is not None else '',
            data.get('motivo_parada', ''),
            data.get('solucao_problema', ''),
            data.get('observacoes', ''),
            'SIM' if sig_tecnico_path else 'NÃO',
            data.get('nome_solicitante', ''),
            'SIM' if sig_solicitante_path else 'NÃO',
            status,
            materiais_str,
            len(nomes_fotos)
        ]

        salvar_registro_excel(row_data, next_row, record_id, total_horas)

        # Salvar fotos no mesmo arquivo Excel (aba "Fotos")
        if fotos_b64:
            salvar_fotos_excel(record_id, fotos_b64)

        # Salvar no Plano de Ação se houver pendências
        if tem_plano_acao:
            salvar_plano_acao(record_id, data, outros_problemas, pecas_necessarias)

        if record_id % 10 == 0:
            fazer_backup()

        horas_fmt = f"{int(total_horas)}h {int((total_horas % 1) * 60)}min" if total_horas else "N/A"
        print(f"  ✓ Registro #{record_id:04d} salvo | Status: {status}")

        return jsonify({
            'success': True,
            'message': f'Registro #{record_id:04d} salvo com sucesso!',
            'id': record_id,
            'total_horas': horas_fmt,
            'status': status,
            'tem_plano_acao': tem_plano_acao
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Erro ao salvar: {str(e)}'}), 500


@app.route('/dashboard')
@login_required
def dashboard():
    try:
        wb = openpyxl.load_workbook(DB_FILE)
        ws = wb["Registros"]
        headers = [cell.value for cell in ws[1]]
        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                # Remove colunas sem cabeçalho (None) para evitar TypeError em sorts/tojson
                record = {k: v for k, v in zip(headers, row) if k is not None}
                records.append(record)
        wb.close()

        total = len(records)
        total_horas = sum(r.get('Total Horas Paradas', 0) or 0 for r in records)
        horas_list = [r.get('Total Horas Paradas', 0) or 0 for r in records if r.get('Total Horas Paradas')]
        tempo_medio = round(sum(horas_list) / len(horas_list), 1) if horas_list else 0

        # Status counts
        finalizadas = sum(1 for r in records if str(r.get('Status', '')).strip() == 'Concluído')
        em_aberto   = sum(1 for r in records if str(r.get('Status', '')).strip() == 'Pendente')

        por_setor = {}
        for r in records:
            setor = r.get('Setor Produtivo', 'N/A') or 'N/A'
            por_setor[setor] = por_setor.get(setor, 0) + 1

        por_motivo = {}
        for r in records:
            motivo = r.get('Motivo da Parada', 'N/A') or 'N/A'
            motivo_curto = motivo.split(' - ')[0] if ' - ' in motivo else motivo
            por_motivo[motivo_curto] = por_motivo.get(motivo_curto, 0) + 1

        # Tipo de manutenção
        por_tipo = {'Corretiva': 0, 'Preventiva': 0, 'Preditiva': 0, 'Melhoria': 0}
        for r in records:
            tipo = str(r.get('Tipo Manutenção', r.get('Motivo da Parada', '')) or '')
            for t in por_tipo:
                if t.lower() in tipo.lower():
                    por_tipo[t] += 1
                    break
            else:
                # Inferir pelo motivo
                motivo = str(r.get('Motivo da Parada', '') or '')
                if 'Preventiva' in motivo or 'Lubrificação' in motivo or '5S' in motivo:
                    por_tipo['Preventiva'] += 1
                elif 'Melhoria' in motivo:
                    por_tipo['Melhoria'] += 1
                elif 'Preditiva' in motivo:
                    por_tipo['Preditiva'] += 1
                else:
                    por_tipo['Corretiva'] += 1

        # Top equipamentos por ocorrências
        por_equipamento = {}
        for r in records:
            eq = r.get('Equipamento') or r.get('Código do Equipamento') or 'N/A'
            por_equipamento[str(eq)] = por_equipamento.get(str(eq), 0) + 1
        top_equipamentos = sorted(por_equipamento.items(), key=lambda x: x[1], reverse=True)[:10]

        # Por mês (últimos 12 meses)
        from collections import defaultdict
        por_mes = defaultdict(int)
        for r in records:
            data_str = str(r.get('Data/Hora Registro') or r.get('Data Ocorrência') or '')
            try:
                if data_str:
                    partes = data_str.split('/')
                    if len(partes) >= 3:
                        mes  = partes[1].zfill(2)
                        ano  = partes[2][:4]
                        if mes.isdigit() and ano.isdigit() and len(ano) == 4:
                            mes_ano = f"{mes}/{ano}"
                            por_mes[mes_ano] += 1
            except Exception:
                pass
        # Ordenar por ano/mês e pegar últimos 12
        meses_ord = sorted(
            (m for m in por_mes.keys() if '/' in m and len(m.split('/')) == 2),
            key=lambda x: (x.split('/')[1], x.split('/')[0])
        )[-12:]
        por_mes_lista = [{'mes': m, 'qtd': por_mes[m]} for m in meses_ord]

        # Status por equipamento para o gráfico de barras
        status_eq = get_status_por_equipamento()
        equipamentos_info = carregar_equipamentos()
        eq_map = {eq['tag']: eq.get('descricao', eq['tag']) for eq in equipamentos_info if eq.get('tag')}

        # Técnico responsável pela OS pendente mais recente por equipamento
        tec_pendente_eq = {}
        for r in records:
            if str(r.get('Status', '')).strip() == 'Pendente':
                tag = str(r.get('Código do Equipamento') or r.get('Equipamento') or '').strip()
                if tag:
                    tec_pendente_eq[tag] = str(r.get('Técnico Responsável') or '').strip()

        grafico_maquinas = []
        for tag, status in status_eq.items():
            grafico_maquinas.append({
                'tag': tag,
                'descricao': eq_map.get(tag, tag),
                'status': status,
                'tecnico_pendente': tec_pendente_eq.get(tag, '')
            })
        grafico_maquinas.sort(key=lambda x: (x['status'] == 'Sem registros', x['status'] == 'Concluído', str(x['tag'] or '')))

        n_pendente  = sum(1 for m in grafico_maquinas if m['status'] == 'Pendente')
        n_ok        = sum(1 for m in grafico_maquinas if m['status'] == 'Concluído')
        n_sem_dados = sum(1 for m in grafico_maquinas if m['status'] == 'Sem registros')

        # Dados do Plano de Ação para drill-down
        plano_records = []
        if PLANO_ACAO_FILE.exists():
            try:
                wb_p = openpyxl.load_workbook(PLANO_ACAO_FILE, read_only=True, data_only=True)
                ws_p = wb_p.worksheets[0]
                h_p = [str(c.value or '').strip() for c in next(ws_p.iter_rows(max_row=1))]
                for row in ws_p.iter_rows(min_row=2, values_only=True):
                    if any(v is not None for v in row):
                        plano_records.append({k: v for k, v in zip(h_p, row) if k})
                wb_p.close()
            except Exception:
                pass

        # Status dos técnicos: 3 estados (Em Atendimento / Form Aberto / Disponível)
        forms_abertos = _get_forms_abertos()
        tecs_pendente = set()
        for r in records:
            if str(r.get('Status', '')).strip() == 'Pendente':
                tec = str(r.get('Técnico Responsável') or '').strip()
                if tec:
                    tecs_pendente.add(tec)
        todos_tecs = sorted(
            {str(r.get('Técnico Responsável') or '').strip() for r in records if r.get('Técnico Responsável')}
            | forms_abertos
        )
        tec_status = []
        for t in todos_tecs:
            if t in tecs_pendente:
                status = 'Em Atendimento'
            elif t in forms_abertos:
                status = 'Form Aberto'
            else:
                status = 'Disponível'
            tec_status.append({'nome': t, 'status': status})

        # Anos disponíveis para filtro
        anos_set = set()
        for r in records:
            data_raw = str(r.get('Data/Hora Registro', '') or '')
            partes = data_raw.split('/')
            if len(partes) >= 3:
                ano = partes[2][:4]
                if ano.isdigit() and len(ano) == 4:
                    anos_set.add(ano)
        anos = sorted(anos_set, reverse=True) or [str(datetime.now().year)]

        return render_template(
            'dashboard.html',
            records=records[-50:],
            records_json=records,
            plano_json=plano_records,
            tec_status=tec_status,
            total=total,
            finalizadas=finalizadas,
            em_aberto=em_aberto,
            tempo_medio=tempo_medio,
            total_horas=round(total_horas, 1),
            por_setor=por_setor,
            por_motivo=por_motivo,
            por_tipo=por_tipo,
            top_equipamentos=top_equipamentos,
            por_mes_lista=por_mes_lista,
            grafico_maquinas=grafico_maquinas,
            n_pendente=n_pendente,
            n_ok=n_ok,
            n_sem_dados=n_sem_dados,
            anos=anos
        )
    except Exception as e:
        import traceback as _tb
        _full = _tb.format_exc()
        return f"<pre style='font-size:13px;padding:20px'><b>Erro dashboard:</b>\n{_full}</pre>", 500


@app.route('/historico')
@login_required
def historico():
    """Página de histórico de equipamentos."""
    tag_busca = request.args.get('tag', '').strip().upper()
    desc_busca = request.args.get('desc', '').strip().lower()
    equipamentos_info = carregar_equipamentos()

    registros = []
    if tag_busca or desc_busca:
        try:
            wb = openpyxl.load_workbook(DB_FILE)
            ws = wb["Registros"]
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(v is not None for v in row):
                    r = dict(zip(headers, row))
                    eq_tag  = str(r.get('Código do Equipamento') or '').upper()
                    eq_desc = str(r.get('Equipamento') or '').lower()
                    if tag_busca and tag_busca in eq_tag:
                        registros.append(r)
                    elif desc_busca and desc_busca in eq_desc:
                        registros.append(r)
            wb.close()
        except Exception:
            pass

    registros_rev = list(reversed(registros))
    total_h = sum(float(r.get('Total Horas Paradas') or 0) for r in registros)
    fins    = sum(1 for r in registros if str(r.get('Status','')).strip() == 'Concluído')
    pends   = sum(1 for r in registros if str(r.get('Status','')).strip() == 'Pendente')
    med_h   = round(total_h / len(registros), 1) if registros else 0

    return render_template(
        'historico.html',
        registros=registros_rev,
        total=len(registros),
        total_horas=round(total_h, 1),
        finalizadas=fins,
        pendentes=pends,
        media_horas=med_h,
        tag_busca=tag_busca,
        desc_busca=desc_busca,
        equipamentos=equipamentos_info
    )


@app.route('/api/historico/<tag>')
@login_required
def api_historico_tag(tag):
    """Retorna os últimos registros de um equipamento (para painel inline no formulário)."""
    tag_upper = tag.strip().upper()
    registros = []
    try:
        wb = openpyxl.load_workbook(DB_FILE)
        ws = wb["Registros"]
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                r = dict(zip(headers, row))
                eq_tag = str(r.get('Código do Equipamento') or '').upper()
                if tag_upper and tag_upper in eq_tag:
                    registros.append({
                        'id':        r.get('ID'),
                        'data':      str(r.get('Data/Hora Registro') or ''),
                        'equipamento': str(r.get('Equipamento') or ''),
                        'tag':       eq_tag,
                        'tecnico':   str(r.get('Técnico Responsável') or ''),
                        'motivo':    str(r.get('Motivo da Parada') or ''),
                        'solucao':   str(r.get('Solução do Problema') or ''),
                        'horas':     r.get('Total Horas Paradas'),
                        'status':    str(r.get('Status') or 'Concluído'),
                        'setor':     str(r.get('Setor Produtivo') or ''),
                    })
        wb.close()
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

    registros_rev = list(reversed(registros))
    total_h = sum(float(r.get('horas') or 0) for r in registros_rev)
    return jsonify({
        'success': True,
        'tag': tag_upper,
        'total': len(registros_rev),
        'total_horas': round(total_h, 1),
        'registros': registros_rev[:10]   # últimos 10 para o painel inline
    })


@app.route('/download')
@login_required
def download():
    fazer_backup()
    return send_file(
        DB_FILE,
        as_attachment=True,
        download_name=f"Manutencao_TPM_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )


@app.route('/download/plano-acao')
@login_required
def download_plano_acao():
    if not PLANO_ACAO_FILE.exists():
        criar_excel_plano_acao()
    return send_file(
        PLANO_ACAO_FILE,
        as_attachment=True,
        download_name=f"Plano_de_Acao_{datetime.now().strftime('%Y%m%d')}.xlsx"
    )


@app.route('/api/registros')
@login_required
def api_registros():
    try:
        wb = openpyxl.load_workbook(DB_FILE)
        ws = wb["Registros"]
        headers = [cell.value for cell in ws[1]]
        records = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                records.append(dict(zip(headers, row)))
        wb.close()
        return jsonify({'success': True, 'total': len(records), 'registros': records[-100:]})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================
# FORMULÁRIOS ABERTOS — rastreamento em tempo real
# ============================================================
# Dict em memória: { 'Nome Tecnico': timestamp_last_ping }
_forms_abertos = {}
_FORM_TIMEOUT = 60  # segundos sem ping = form considerado fechado

@app.route('/api/form-ping', methods=['POST'])
@login_required
def api_form_ping():
    """Mobile form faz ping a cada 25s enquanto está aberto."""
    import time
    data = request.get_json(silent=True) or {}
    tecnico = str(data.get('tecnico', '')).strip()
    if tecnico:
        _forms_abertos[tecnico] = time.time()
    return jsonify({'ok': True})

@app.route('/api/form-ping', methods=['DELETE'])
@login_required
def api_form_close():
    """Mobile form notifica que foi fechado/submetido."""
    data = request.get_json(silent=True) or {}
    tecnico = str(data.get('tecnico', '')).strip()
    _forms_abertos.pop(tecnico, None)
    return jsonify({'ok': True})

def _get_forms_abertos():
    """Retorna set de técnicos com form aberto (ping recente)."""
    import time
    agora = time.time()
    return {t for t, ts in list(_forms_abertos.items()) if agora - ts < _FORM_TIMEOUT}

@app.route('/api/dashboard-json')
@login_required
def api_dashboard_json():
    """Retorna dados resumidos do dashboard para polling em tempo real."""
    try:
        records = []
        if DB_FILE.exists():
            wb = openpyxl.load_workbook(DB_FILE, read_only=True, data_only=True)
            if 'Registros' in wb.sheetnames:
                ws = wb['Registros']
                headers = [str(c.value or '').strip() for c in next(ws.iter_rows(max_row=1))]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(v is not None for v in row):
                        records.append({k: v for k, v in zip(headers, row) if k})
            wb.close()

        total = len(records)
        finalizadas = sum(1 for r in records if str(r.get('Status', '')).strip() == 'Concluído')
        em_aberto   = total - finalizadas

        # Status dos técnicos com 3 estados
        forms_abertos = _get_forms_abertos()
        tecs_pendente = set()
        for r in records:
            if str(r.get('Status', '')).strip() == 'Pendente':
                t = str(r.get('Técnico Responsável') or '').strip()
                if t:
                    tecs_pendente.add(t)

        todos_tecs = sorted(
            {str(r.get('Técnico Responsável') or '').strip() for r in records if r.get('Técnico Responsável')}
            | forms_abertos
        )

        tec_status = []
        for t in todos_tecs:
            if t in tecs_pendente:
                status = 'Em Atendimento'
            elif t in forms_abertos:
                status = 'Form Aberto'
            else:
                status = 'Disponível'
            tec_status.append({'nome': t, 'status': status})

        return jsonify({
            'total': total,
            'finalizadas': finalizadas,
            'em_aberto': em_aberto,
            'tec_status': tec_status,
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================
# ROTAS: PAINEL DE ADMINISTRAÇÃO
# ============================================================
@app.route('/admin')
@admin_required
def admin():
    usuarios = carregar_usuarios()
    listas = carregar_listas()
    equipamentos = carregar_equipamentos()
    producao = carregar_producao()
    estoque = carregar_estoque()
    # Índice de producao por tag para fácil lookup no template
    producao_por_tag = {p['tag']: p for p in producao}
    # Colaboradores: do Colaboradores.xlsx + técnicos sem cadastro detalhado
    colaboradores = carregar_colaboradores()
    nomes_cad = {c.get('Nome', '').strip().lower() for c in colaboradores}
    for t in listas.get('tecnicos', []):
        tn = str(t).strip()
        if tn and tn.lower() not in nomes_cad and not tn.lower().startswith('outro'):
            colaboradores.append({'Nome': tn})
    return render_template('admin.html',
        usuarios=usuarios,
        listas=listas,
        equipamentos=equipamentos,
        equipamentos_lista=equipamentos,
        producao=producao,
        producao_por_tag=producao_por_tag,
        estoque=estoque,
        colaboradores=colaboradores,
        funcoes=FUNCOES_PADRAO
    )


# --- Usuários ---
@app.route('/admin/usuarios/adicionar', methods=['POST'])
@admin_required
def admin_adicionar_usuario():
    try:
        dados = request.get_json()
        usuarios = carregar_usuarios()

        # Verificar duplicidade de username
        if any(u['username'] == dados['username'] for u in usuarios):
            return jsonify({'success': False, 'message': 'Nome de usuário já existe.'})

        novo_id = max((u['id'] for u in usuarios), default=0) + 1
        novo = {
            'id': novo_id,
            'username': dados['username'].strip(),
            'nome': dados['nome'].strip(),
            'password_hash': hash_senha(dados['senha']),
            'perfil': dados['perfil'],
            'ativo': True,
            'criado_em': datetime.now().strftime('%d/%m/%Y %H:%M:%S')
        }
        usuarios.append(novo)
        salvar_usuarios(usuarios)
        return jsonify({'success': True, 'message': f'Usuário "{novo["nome"]}" criado com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/admin/usuarios/editar/<int:uid>', methods=['POST'])
@admin_required
def admin_editar_usuario(uid):
    try:
        dados = request.get_json()
        usuarios = carregar_usuarios()
        atual = usuario_logado()

        for u in usuarios:
            if u['id'] == uid:
                u['nome'] = dados['nome'].strip()
                u['perfil'] = dados['perfil']
                u['ativo'] = dados.get('ativo', True)
                if dados.get('senha'):
                    u['password_hash'] = hash_senha(dados['senha'])
                break
        else:
            return jsonify({'success': False, 'message': 'Usuário não encontrado.'})

        salvar_usuarios(usuarios)
        return jsonify({'success': True, 'message': 'Usuário atualizado com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/admin/usuarios/excluir/<int:uid>', methods=['POST'])
@admin_required
def admin_excluir_usuario(uid):
    try:
        atual = usuario_logado()
        if atual['id'] == uid:
            return jsonify({'success': False, 'message': 'Você não pode excluir seu próprio usuário.'})

        usuarios = carregar_usuarios()
        admins = [u for u in usuarios if u['perfil'] == 'admin' and u.get('ativo', True)]
        alvo = next((u for u in usuarios if u['id'] == uid), None)

        if not alvo:
            return jsonify({'success': False, 'message': 'Usuário não encontrado.'})
        if alvo['perfil'] == 'admin' and len(admins) <= 1:
            return jsonify({'success': False, 'message': 'Não é possível excluir o único administrador.'})

        usuarios = [u for u in usuarios if u['id'] != uid]
        salvar_usuarios(usuarios)
        return jsonify({'success': True, 'message': 'Usuário excluído com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# --- Listas (Técnicos e Setores) ---
@app.route('/admin/listas/salvar', methods=['POST'])
@admin_required
def admin_salvar_listas():
    try:
        dados = request.get_json()
        listas = carregar_listas()

        if 'tecnicos' in dados:
            tecnicos = [t.strip() for t in dados['tecnicos'] if t.strip()]
            listas['tecnicos'] = tecnicos

        if 'setores' in dados:
            setores = [s.strip() for s in dados['setores'] if s.strip()]
            listas['setores'] = setores

        if 'motivos' in dados:
            motivos = [m.strip() for m in dados['motivos'] if m.strip()]
            listas['motivos'] = motivos

        if 'solucoes' in dados:
            solucoes = [s.strip() for s in dados['solucoes'] if s.strip()]
            listas['solucoes'] = solucoes

        salvar_listas(listas)
        return jsonify({'success': True, 'message': 'Listas atualizadas com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/admin/colaboradores/salvar', methods=['POST'])
@admin_required
def admin_salvar_colaboradores():
    try:
        dados = request.get_json()
        lista = dados.get('colaboradores', [])
        norm, nomes = [], []
        for c in lista:
            nome = str(c.get('Nome', '') or '').strip()
            if not nome:
                continue
            norm.append({h: str(c.get(h, '') or '').strip() for h in COLAB_HEADERS})
            nomes.append(nome)
        salvar_colaboradores(norm)
        # mantém os nomes sincronizados na lista de técnicos do formulário de OS
        listas = carregar_listas()
        listas['tecnicos'] = nomes
        salvar_listas(listas)
        return jsonify({'success': True,
                        'message': f'{len(norm)} colaborador(es) salvo(s) no Colaboradores.xlsx!'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


@app.route('/admin/equipamentos/salvar', methods=['POST'])
@admin_required
def admin_salvar_equipamentos():
    try:
        dados = request.get_json()
        lista = dados.get('equipamentos', [])
        equipamentos = []
        tags_vistas = set()
        for eq in lista:
            descricao = eq.get('descricao', '').strip()
            tag = eq.get('tag', '').strip().upper()
            area = eq.get('area', '').strip()
            fornecedor = eq.get('fornecedor', '').strip()
            if not descricao and not tag:
                continue
            if tag and tag in tags_vistas:
                return jsonify({'success': False,
                                'message': f'TAG duplicada: {tag}. Cada equipamento precisa de uma TAG única.'}), 400
            if tag:
                tags_vistas.add(tag)
            equipamentos.append({'descricao': descricao, 'tag': tag,
                                  'area': area, 'fornecedor': fornecedor})
        salvar_equipamentos(equipamentos)
        return jsonify({'success': True,
                        'message': f'{len(equipamentos)} equipamento(s) salvos com sucesso!'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================
# FUNCOES AUXILIARES
# ============================================================
def calcular_horas(inicio, fim):
    try:
        if not inicio or not fim:
            return None
        fmt = '%H:%M'
        t_inicio = datetime.strptime(inicio, fmt)
        t_fim = datetime.strptime(fim, fmt)
        delta = t_fim - t_inicio
        if delta.total_seconds() < 0:
            delta += timedelta(days=1)
        return delta.total_seconds() / 3600
    except Exception:
        return None


def salvar_assinatura(base64_data, nome_arquivo):
    try:
        if not base64_data or base64_data == 'data:,':
            return None
        if ',' in base64_data:
            base64_data = base64_data.split(',', 1)[1]
        if not base64_data.strip():
            return None
        img_data = base64.b64decode(base64_data)
        filepath = ASSINATURAS_DIR / f"{nome_arquivo}.png"
        with open(filepath, 'wb') as f:
            f.write(img_data)
        return str(filepath)
    except Exception:
        return None


def obter_proximo_id():
    wb = openpyxl.load_workbook(DB_FILE)
    ws = wb["Registros"]
    ultima_linha = 1
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(v is not None for v in row):
            ultima_linha += 1
    wb.close()
    return ultima_linha, ultima_linha + 1


def salvar_registro_excel(row_data, next_row, record_id, total_horas):
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border_thin = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    fill_color = "EBF1F8" if record_id % 2 == 0 else "FFFFFF"
    row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

    max_tentativas = 5
    ultimo_erro = None

    for tentativa in range(1, max_tentativas + 1):
        temp_path = DB_FILE.parent / f"_temp_registro_{record_id}.xlsx"
        try:
            wb = openpyxl.load_workbook(DB_FILE)
            ws = wb["Registros"]
            linha_real = ws.max_row + 1

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=linha_real, column=col, value=value)
                cell.fill = row_fill
                cell.alignment = align_center
                cell.border = border_thin

            wb.save(temp_path)
            wb.close()
            os.replace(temp_path, DB_FILE)
            return True

        except PermissionError as e:
            ultimo_erro = e
            if tentativa < max_tentativas:
                time.sleep(1.5)
            if temp_path.exists():
                try:
                    temp_path.unlink()
                except Exception:
                    pass

        except Exception as e:
            ultimo_erro = e
            if temp_path.exists():
                try:
                    temp_path.unlink()
                except Exception:
                    pass
            break

    print(f"[ERRO] salvar_registro_excel falhou apos {max_tentativas} tentativas: {ultimo_erro}")
    return False


@app.route('/plano-acao')
@login_required
def plano_acao():
    registros = []
    if PLANO_ACAO_FILE.exists():
        try:
            wb = openpyxl.load_workbook(PLANO_ACAO_FILE, read_only=True)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(v is not None for v in row):
                    registros.append(dict(zip(headers, row)))
            wb.close()
        except Exception:
            pass
    return render_template('plano_acao.html', registros=registros)


@app.route('/admin/producao/salvar', methods=['POST'])
@admin_required
def admin_salvar_producao():
    try:
        dados = request.get_json()
        lista = dados.get('producao', [])
        resultado = []
        for item in lista:
            tag = item.get('tag', '').strip().upper()
            if not tag:
                continue
            modelos = []
            for m in item.get('modelos', []):
                nome = m.get('nome', '').strip()
                try:
                    volume = float(m.get('volume', 0) or 0)
                except (ValueError, TypeError):
                    volume = 0
                if nome:
                    modelos.append({'nome': nome, 'volume': volume})
            try:
                indice_qualidade = float(item.get('indice_qualidade', 100) or 100)
            except (ValueError, TypeError):
                indice_qualidade = 100
            try:
                tempo_turno = float(item.get('tempo_turno', 8) or 8)
            except (ValueError, TypeError):
                tempo_turno = 8
            try:
                tempo_mes = float(item.get('tempo_mes', 176) or 176)
            except (ValueError, TypeError):
                tempo_mes = 176
            resultado.append({
                'tag': tag,
                'modelos': modelos,
                'indice_qualidade': indice_qualidade,
                'tempo_turno': tempo_turno,
                'tempo_mes': tempo_mes
            })
        salvar_producao(resultado)
        return jsonify({'success': True, 'message': f'{len(resultado)} equipamento(s) configurados!'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================
# ROTAS: BACKUP / DOWNLOAD
# ============================================================
@app.route('/admin/backup/download/manutencao')
@admin_required
def backup_manutencao():
    if not DB_FILE.exists():
        return "Arquivo não encontrado no servidor.", 404
    return send_file(
        str(DB_FILE),
        as_attachment=True,
        download_name=f"Manutencao_TPM_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/admin/backup/download/plano')
@admin_required
def backup_plano():
    if not PLANO_ACAO_FILE.exists():
        return "Arquivo não encontrado no servidor.", 404
    return send_file(
        str(PLANO_ACAO_FILE),
        as_attachment=True,
        download_name=f"Plano_de_Acao_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/admin/backup/download/config')
@admin_required
def backup_config():
    import zipfile, io
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for f in CONFIG_DIR.glob('*.json'):
            zf.write(f, f.name)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=f"Config_TPM_{datetime.now().strftime('%Y%m%d_%H%M')}.zip",
        mimetype='application/zip'
    )


# ============================================================
# API DE LEITURA (consumida pelo portal SGM) — protegida por chave
# ============================================================
SGM_API_KEY = os.environ.get('SGM_API_KEY', 'caloi-sgm-2026')

@app.route('/api/dados/<dataset>')
def api_dados(dataset):
    if request.args.get('key', '') != SGM_API_KEY:
        return jsonify({'error': 'unauthorized'}), 401
    mapa = {'os': DB_FILE, 'plano': PLANO_ACAO_FILE, 'colaboradores': COLAB_FILE}
    path = mapa.get(dataset)
    if not path or not path.exists():
        return jsonify([])
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        linhas = list(ws.iter_rows(values_only=True))
        wb.close()
        if not linhas:
            return jsonify([])
        headers = [str(h).strip() if h is not None else '' for h in linhas[0]]
        out = []
        for r in linhas[1:]:
            if any(v is not None for v in r):
                out.append({headers[i]: (None if i >= len(r) or r[i] is None else str(r[i]))
                            for i in range(len(headers))})
        return jsonify(out)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# Inicializar banco e config sempre que o processo sobe (local + nuvem)
init_database()
init_config()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_DEBUG', 'false').lower() == 'true'
    print(f"=== Caloi TPM v2.0 === porta {port}")
    app.run(debug=debug, host='0.0.0.0', port=port)
